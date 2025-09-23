import sys
from openpyxl import load_workbook
wb = load_workbook("fields table.xlsx", data_only=True)
ws = wb.active
rows = list(ws.iter_rows(values_only=True))
for r in rows:
    if r is None:
        continue
    vals = ["" if v is None else str(v).replace('\t',' ').replace('\n',' ') for v in r]
    print('\t'.join(vals))
